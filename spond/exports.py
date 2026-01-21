"""Utilities for reading downloaded spreadsheet exports."""

from pathlib import Path
import openpyxl as op
from typing import NamedTuple, Optional, Any, Iterable, Iterator
from collections import deque


class User(NamedTuple):
    """A user's basic information."""

    name: str
    email: Optional[str]
    phone: Optional[str]


def _str_or_none(v: Optional[Any]) -> Optional[str]:
    if v is None or v == "":
        return None
    return str(v)


def read_poll(fpath: Path) -> Iterator[tuple[User, Optional[set[str]]]]:
    """Read an exported poll result excel sheet.

    Parameters
    ----------
    fpath
        Path to .xlsx

    Yields
    ------
    tuple[User, set[str] | None]
        Users and which options they voted for.
        Empty set if they voted blank;
        None if they did not vote.
    """
    wb: op.Workbook = op.load_workbook(fpath)
    sheet = wb[wb.sheetnames[0]]
    block_n = 0
    rows_iter = sheet.iter_rows()
    for row in rows_iter:
        val = row[0].value
        if _str_or_none(val) is None:
            block_n += 1
            if block_n >= 3:
                break

    # name, email, phone
    values = [h.value for h in next(rows_iter)[:-3]]

    for row in rows_iter:
        this_row = [c.value for c in row]
        phone = _str_or_none(this_row.pop())
        email = _str_or_none(this_row.pop())
        name = str(this_row.pop())
        user = User(name, email, phone)

        responses = set()

        # voted blank
        if this_row.pop():
            yield (user, responses)
            continue

        for val, response in zip(values, this_row):
            if response:
                responses.add(val)

        yield (User(name, email, phone), responses or None)


class UserExt(NamedTuple):
    """A user, extended information, and their group memberships"""

    user: User
    info: dict[str, Any]
    groups: set[str]


def _sliding_window(seq: Iterable, n: int) -> Iterator[tuple]:
    d = deque(maxlen=n)
    for val in seq:
        d.append(val)
        if len(d) >= n:
            yield tuple(d)


def read_members(fpath: Path) -> Iterator[UserExt]:
    """Read an exported membership list spreadsheet.

    Parameters
    ----------
    fpath
        Path to .xlsx file.

    Yields
    ------
    UserExt
        Extended user information

    Raises
    ------
    RuntimeError
        Could not find expected basic information columns
    """
    wb: op.Workbook = op.load_workbook(fpath)
    sheet = wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter)
    search_titles = ("Name", "Email", "Cell")
    for user_idx, titles_tup in enumerate(
        _sliding_window(header_row, len(search_titles))
    ):
        if titles_tup != search_titles:
            continue
        groups = [str(h) for h in header_row[:user_idx]]
        info_keys = [str(h) for h in header_row[user_idx + 3 :]]
        break
    else:
        raise RuntimeError("Name, Email, Cell columns not found")

    for row in rows_iter:
        cell_iter = iter(row)
        grps = {name for name, cell in zip(groups, cell_iter) if _str_or_none(cell)}
        u = User(
            str(next(cell_iter)),
            _str_or_none(next(cell_iter)),
            _str_or_none(next(cell_iter)),
        )
        info = dict(zip(info_keys, cell_iter))
        yield UserExt(u, info, grps)
